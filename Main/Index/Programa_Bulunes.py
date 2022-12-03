#*********************************************************
# Primer commit main app-prueba / Programa: Bulones0A
# *Subir a Github
# *Index.py
# ejecutable.exe
# release: 31/10/2022
# Developer: H. Leandro Mangicavalli
#*********************************************************

from ast import While
#from cgitb import reset
from operator import concat
from tkinter import N, Y
import Calendario
# libreria para excel
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook

# Leer el archivo
wb = openpyxl.load_workbook("C:\DESARROLLO\Bulones0A\Main\Index\Tabla_bulones.xlsx", data_only=True)
#wb = openpyxl.load_workbook("https://github.com/mangicl/Bulones0A/blob/main/Main/Index/Tabla_bulones.xlsx")
# Fijar la hoja
ws1 = wb["Stock"]
celdas1 = ws1["B3" : "F106"]
lista_bulones = []
ws2 = wb["Clientes"]
celdas2 = ws2["B2" : "D51"]
lista_clientes = []

#main - resources
deseaComenzar = Y,N
deseaContinuar = Y,N
seleccion = 0
#diametro_stock = 6, 7, 8, 9, 10, 16, 18, 20, 22, 24, 27, 30, 33, 36, 39, 42, 45, 52, 56, 60, 64, 68, 72, 76, 80
diametro_stock = 0
#paso_stock = 1, 1.25, 1.5, 2, 2.25, 3, 3.5, 4, 4.5, 5, 5.5, 6
paso_stock = 0
#largo_stock = 1/4, 1/2, 3/4, 1
largo_stock = 0
rosca_izq = "izquierda"
rosca_der = "derecha"
unidad1 = "unidad"
unidad2 = "kilo"

#Variable de contabilidad
cantUnidad = 0
cuentaUnidad = 0
cantKilos = 0
cuentaKilos = 0

while True:
    deseaComenzar = input("¿Desea comenzar/continuar? (Y para SI / N para NO): ")
    if deseaComenzar == Y:
        print("")
        print("Ingrese las datos a continuación: ")
        print("")
    elif deseaComenzar == N:
        print("")
        print("¡Muchas gracias por su compra!")
        print("   Que tenga un buen día.   ")
        print("")
        break
    
    while True:
        if seleccion != 6:
            print("""Seleccione el tipo de operacion:

                1) Consultar stock disponible
                2) Seleccionar bulon a comprar
                3) Ver totales de la compra
                4) Seguir comprando
                5) Borrar datos ingresados y volver a comenzar
                6) Salir
            """)
            seleccion = int(input("Elija una opcion: "))

            if seleccion == 1:
                print(" ")
                print("Usted ha elegido consultar el stock disponible")
                print(" ")
                for fila in celdas1:
                    bulones = [celda.value for celda in fila]
                    lista_bulones.append(bulones)
#                   print([celda.value for celda in fila])
                for bulones in lista_bulones:
                    print(f"""Bulones de diametro {bulones[0]} con paso {bulones[1]} y largo {bulones[2]} hay en stock:
                    {bulones[3]} con rosca izquierda y {bulones[4]} con rosca derecha""")
                break

            if seleccion == 2:
                print(" ")
                print("Comencemos a ingresar los datos del bulon...")
                print(" ")
                cantUnidad = 0
                cantKilos = 0

                while True:
                    diametro_input = int(input("Ingrese el diametro en milimetros: "))
                    if diametro_input != 6 and diametro_input != 7 and diametro_input != 8 and diametro_input != 9 and diametro_input != 10 and diametro_input != 16 and diametro_input != 18 and diametro_input != 20 and diametro_input != 22 and diametro_input != 24 and diametro_input != 27 and diametro_input != 30 and diametro_input != 33 and diametro_input != 36 and diametro_input != 39 and diametro_input != 42 and  diametro_input != 45 and diametro_input != 48 and diametro_input != 52 and diametro_input != 56 and diametro_input != 60 and diametro_input != 64 and diametro_input != 68 and diametro_input != 72 and diametro_input != 76 and diametro_input != 80:
                        print("Introduzca un diametro valido")
                    else:
                        print("Diametro valido")
                        break
                    continue

                while True:
                    paso_input = input("Ingrese el paso: ")
                    if paso_input != "1" and paso_input != "1.25" and paso_input != "1.5" and paso_input != "2" and paso_input != "2.25" and paso_input != "3" and paso_input != "3.5" and paso_input != "4" and paso_input != "4.5" and paso_input != "5" and paso_input != "5.5" and paso_input != "6":
                        print("Introduzca un paso valido")
                    else:
                        print("Paso valido")
                        break
                    continue

                while True:
                    largo_input = input("Ingrese el largo en pulgadas: ")
                    if largo_input != "1/4" and largo_input != "1/2" and largo_input != "3/4" and largo_input != "1":
                        print("Introduzca un largo valido")
                    else:
                        print("Largo valido")
                        break
                    continue

                while True:
                    rosca_input = input("Ingrese el tipo de rosca (IZQUIERDA / DERECHA): ")
                    if rosca_input != rosca_izq and rosca_input != rosca_der:
                        print("Introduzca un tipo de rosca valido")
                    else:
                        print("Rosca valida")
                        break
                    continue

                while True:
                    unidad_input = input("Indique si compra por UNIDAD o por KILO): ")
                    if unidad_input != unidad1 and unidad_input != unidad2:
                        print("Indique una unidad valida")
                    elif unidad_input == unidad1:
                        cantUnidad = int(input("Ingrese la cantidad de unidades requeridas: "))
                        break
                    elif unidad_input == unidad2:
                        cantKilos = float(input("Ingrese la cantidad de kilos requeridos (10000 unidades = 1 Kg): "))
                        break
                    continue
                    
                while True:
                    if unidad_input == unidad1:
                        print(f"Usted ha elegido {cantUnidad} bulones con diametro: {diametro_input}mm / paso: {paso_input} / largo {largo_input} de pulgada / rosca {rosca_input}")
                        break                   
                    if unidad_input == unidad2:    
                        print(f"Usted ha elegido {cantKilos}Kg de bulones con diametro: {diametro_input}mm / paso: {paso_input} / largo {largo_input} de pulgada / rosca {rosca_input}")
                        break

                while True:
                    deseaContinuar = input("Confirmar si es correcto (Y para SI / N para NO): ")
                    if deseaContinuar == Y:
                        cuentaUnidad += cantUnidad
                        cuentaKilos += cantKilos
                        print("Excelente")
                        break    
                    elif deseaContinuar == N:
                        print("¡Volvamos a empezar!")
                        break
                    continue             
                    
            elif seleccion == 3:
                print(f"""Usted esta comprando...
                            Bulones por unidad: {cuentaUnidad}
                            Bulones por kilo: {cuentaKilos}
                 """)

            elif seleccion == 4:
                print(" ")
                print("Usted ha elegido seguir comprando")
                print(" ")

            elif seleccion == 5:
                cantUnidad = 0
                cuentaUnidad = 0
                cantKilos = 0
                cuentaKilos = 0
                print(" ")
                print("Se han borrado todos los datos ingresados")
                print(" ")

            elif seleccion == 6:
                print(" ")
                print("Ha salido de la compra con exito")
                print(" ")
                break
            