#******************************************
# Primer commit main app-prueba
# subir a Github
# index
# ejecutable.exe
# release: 31/10/2022
# Developer: H. Leandro Mangicavalli
#******************************************

from ast import While
#from cgitb import reset
from operator import concat
from tkinter import N, Y
import Calendario
import pandas as pd
import openpyxl
from openpyxl import load_workbook

# datos = pd.read_csv("data_base.csv")
# print(datos[[""]])
#wb = "F:\CURSOS\Curso Python\Tabla bulones.xlsx"

#datos = pd.read_excel("F:\CURSOS\Curso Python\Tabla bulones.xlsx")
#print(datos)

#datos = openpyxl.load_workbook("F:\CURSOS\Curso Python\Tabla bulones.xlsx")
#print(datos)

wb = "C:\DESARROLLO\Bulones0A\Main\Index\Tabla_bulones.xlsx"

#datos = openpyxl.load_workbook("C:\DESARROLLO\Bulones0A\Main\Index\Tabla_bulones.xlsx")
datos2 = pd.read_excel("C:\DESARROLLO\Bulones0A\Main\Index\Tabla_bulones.xlsx")
#print(datos)
print(datos2)


#main - resources
deseaComenzar = Y,N
continuar = Y,N
seleccion = 0
diametro = [6, 7, 8, 9, 10, 16, 18, 20, 22, 24, 27, 30, 33, 36, 39, 42, 45, 52, 56, 60, 64, 68, 72, 76, 80]
paso = 1, 1.25, 1.5, 2, 2.25, 3, 3.5, 4, 4.5, 5, 5.5, 6
largo = 1/4, 1/2, 3/4, 1
rosca = "izquierda", "derecha"
cantidad = "unidad", "kilo"

#Variable de contabilidad
cuentaCantidad = 0
cuentaCantidadTotal = cuentaCantidad
cuentaKilos = 0
cuentaKilosTotal = cuentaKilos

#index

#diametro = input ("Indique diametro")
#print("usted a ingresado: " + diametro)

#paso = input ("Indique paso")
#print("usted a ingresado: " + paso)

#largo = input ("Indique largo")
#print("usted a ingresado: " + largo)

#rosca = input ("Indique rosca")
#print("usted a ingresado: " + rosca)

#cantidad = input ("Indique cantidad")
#print("usted a ingresado: " + cantidad)

<<<<<<< HEAD
#def deseaComenzar ():

=======
>>>>>>> 5413ca7cc1105352dcb143c2eac3798aebbf718a
while True:
    deseaComenzar = input("¿Desea comenzar? (Y para SI / N para NO)")
    if deseaComenzar != N:
        print("Ingrese las datos a continuación:")
    elif deseaComenzar == N:
        print("")
        print("Hasta pronto")
        print("")
        break
    
    while True:
<<<<<<< HEAD
        diametro = input("Ingrese el diametro")
        if diametro != "6":
            print("Introduzca un diametro valido")
        else:
            print("Correcto")
=======
        if seleccion != 5:
            print("""Seleccione el tipo de operacion:
            1) Bulones de rosca derecha
            2) Bulones de rosca izquierda
            3) Seguir comprando
            4) Salir""")
        seleccion = int(input("Elija una opcion: "))
        if seleccion == 1:
            print(" ")
            print("Rosca derecha seleccionada")
            print(" ")

            while True:
                diametro = input("Ingrese el diametro: ")
                if diametro != "6":
                    print("Introduzca un diametro valido")
                else:
                    print("Diametro valido")
                    break
                    continue

            while True:
                paso = input("Ingrese el paso: ")
                if paso != "1":
                    print("Introduzca un paso valido")
                else:
                    print("Paso valido")
                    break
                continue

            while True:
                largo = input("Ingrese el largo: ")
                if largo != "1/4":
                    print("Introduzca un largo valido")
                else:
                    print("Largo valido")
                    break
                continue


        elif seleccion == 2:
            print(" ")
            print("Rosca izquierda seleccionada")
            print(" ")
        elif seleccion == 3:
            print(" ")
            print("Usted ha elegido seguir comprando")
            print(" ")
        elif seleccion == 4:
            print(" ")
            print("Usted ha salido de la compra")
            print(" ")
>>>>>>> 5413ca7cc1105352dcb143c2eac3798aebbf718a
            break

<<<<<<< HEAD
    while True:
        paso = input("Ingrese el paso")
        if paso != "1":
            print("Introduzca un paso valido")
        else:
            print("Correcto")
            break
        continue

    while True:
        largo = input("Ingrese el largo")
        if largo != "1/4":
            print("Introduzca un largo valido")
        else:
            print("Correcto")
            break
        continue

    while True:
        rosca = input("Ingrese el tipo de rosca (izquierda/derecha)")
        if rosca != "derecha":
            print("Introduzca un tipo de rosca valido")
        else:
            print("Correcto")
            break
        continue

    while True:
        cantidad = input("Indique la cantidad (unidad/kilo)")
        if cantidad != "unidad":
            print("Indique una cantidad valida")
        else:
            print("Correcto")
            break
        continue
=======
#    while True:
#        rosca = input("Ingrese el tipo de rosca (izquierda/derecha)")
#        if rosca != "derecha":
#            print("Introduzca un tipo de rosca valido")
#        else:
#            print("Correcto")
#            break
#        continue

        while True:
            cantidad = input("Indique la cantidad (unidad/kilo): ")
            if cantidad != "unidad":
                print("Indique una cantidad valida")
            else:
                print("Correcto")
                break
            continue
>>>>>>> 5413ca7cc1105352dcb143c2eac3798aebbf718a


